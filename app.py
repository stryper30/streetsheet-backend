from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io
import json
from datetime import datetime, date
import copy
 
app = Flask(__name__)
CORS(app)
 
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
 
def parse_time(val):
    """Convert '08:00' to '8am', '16:30' to '4:30pm'"""
    if not val:
        return ''
    try:
        h, m = map(int, val.split(':'))
        ampm = 'pm' if h >= 12 else 'am'
        h12 = h % 12 or 12
        if m == 0:
            return f'{h12}{ampm}'
        return f'{h12}:{str(m).zfill(2)}{ampm}'
    except:
        return val
 
def parse_date_range(date_range_str):
    """Parse '3/14-3/20' into a start date. Assumes current year."""
    try:
        start_part = date_range_str.split('-')[0].strip()
        m, d = map(int, start_part.split('/'))
        year = datetime.now().year
        return datetime(year, m, d)
    except:
        return None
 
def is_snow_emergency(comment):
    if not comment:
        return False
    c = comment.lower()
    return 'snow' in c and 'emergency' in c
 
def yellow_row(ws, row_num):
    """Apply yellow fill to entire row"""
    for col in range(1, 13):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = YELLOW_FILL
 
def copy_cell_style(source_cell, target_cell):
    """Copy border, font, alignment from source to target"""
    if source_cell.border:
        target_cell.border = copy.copy(source_cell.border)
    if source_cell.font:
        target_cell.font = copy.copy(source_cell.font)
    if source_cell.alignment:
        target_cell.alignment = copy.copy(source_cell.alignment)
 
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': 'StreetSheet Backend'})
 
@app.route('/fill-timesheet', methods=['POST'])
def fill_timesheet():
    try:
        # ── Get template file ──
        if 'template' not in request.files:
            return jsonify({'error': 'No template file provided'}), 400
        
        template_file = request.files['template']
        data = json.loads(request.form.get('data', '{}'))
 
        # ── Load template with openpyxl ──
        template_bytes = template_file.read()
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
 
        # ── Get week start date from template J6 ──
        date_range = ws['J6'].value or ''
        week_start = parse_date_range(str(date_range))
 
        # ── Employee Name ──
        name = data.get('name', '')
        if name:
            ws['C4'] = '                   ' + name + ' '
 
        # ── Worked Hours rows 10-24 ──
        days_data = data.get('days', [])
        ot_entries = data.get('otEntries', {})
        
        current_row = 10
        day_offsets = [0, 1, 2, 3, 4, 5, 6]  # Sat=0, Sun=1, Mon=2...Fri=6
 
        for i, day in enumerate(days_data):
            if current_row > 24:
                break
 
            time_in  = day.get('timeIn', '')
            time_out = day.get('timeOut', '')
            hours    = day.get('hours', 0)
            on_call  = day.get('onCall', False)
            thv      = day.get('thv', False)
            comment  = day.get('comment', '').strip()
 
            # Never skip any day — always write the date for all days Sat-Fri
            if week_start:
                from datetime import timedelta
                row_date = week_start + timedelta(days=day_offsets[i])
            else:
                row_date = None
 
            # Always write the date
            if row_date:
                ws.cell(row=current_row, column=1).value = row_date
                ws.cell(row=current_row, column=1).number_format = 'M/D/YYYY'
 
            # Only write other fields if there's actual data
            if not time_in and not time_out and not hours:
                current_row += 1
                continue
            if time_in:
                ws.cell(row=current_row, column=2).value = parse_time(time_in)
            if time_out:
                ws.cell(row=current_row, column=3).value = parse_time(time_out)
            if hours:
                ws.cell(row=current_row, column=4).value = float(hours)
            if on_call:
                ws.cell(row=current_row, column=5).value = 'Yes'
            if thv:
                ws.cell(row=current_row, column=6).value = 'Yes'
            auth_by = day.get('authBy', '').strip().upper()
            if auth_by:
                ws.cell(row=current_row, column=7).value = auth_by
            if comment:
                ws.cell(row=current_row, column=8).value = comment
 
            current_row += 1
 
            # Write OT rows for this day
            day_ot = ot_entries.get(str(i), [])
            for ot in day_ot:
                if current_row > 24:
                    break
                ot_time_in  = ot.get('timeIn', '')
                ot_time_out = ot.get('timeOut', '')
                ot_hours    = ot.get('hours', 0)
                ot_comment  = ot.get('comment', '').strip()
                ot_auth     = ot.get('authBy', '').strip().upper()
 
                if not ot_time_in and not ot_time_out and not ot_hours:
                    continue
 
                if row_date:
                    ws.cell(row=current_row, column=1).value = row_date
                    ws.cell(row=current_row, column=1).number_format = 'M/D/YYYY'
                if ot_time_in:
                    ws.cell(row=current_row, column=2).value = parse_time(ot_time_in)
                if ot_time_out:
                    ws.cell(row=current_row, column=3).value = parse_time(ot_time_out)
                if ot_hours:
                    ws.cell(row=current_row, column=4).value = float(ot_hours)
                if ot_auth:
                    ws.cell(row=current_row, column=7).value = ot_auth
                if ot_comment:
                    ws.cell(row=current_row, column=8).value = ot_comment
 
                # Yellow highlight if snow emergency
                if is_snow_emergency(ot_comment):
                    yellow_row(ws, current_row)
 
                current_row += 1
 
        # ── Pay / Comp row 27 ──
        pay_comp = data.get('payComp', {})
        pay_on   = pay_comp.get('pay', False)
        comp_on  = pay_comp.get('comp', False)
        comp_hrs = pay_comp.get('compHrs', '')
 
        existing = ws['A27'].value or ''
        if pay_on:
            existing = existing.replace('Pay:', 'Pay: Yes')
        if comp_on:
            if comp_hrs:
                existing = existing.replace('Comp:', f'Comp: {comp_hrs} hrs')
            else:
                existing = existing.replace('Comp:', 'Comp: Yes')
        ws['A27'] = existing
 
        # ── Non-Working Hours rows 31-36 ──
        leave_data = data.get('leave', [])
        leave_row = 31
        leave_day_offsets = [2, 3, 4, 5, 6]  # Mon=+2 from Sat
 
        for i, leave in enumerate(leave_data):
            if leave_row > 36:
                break
            l_type  = leave.get('type', '')
            l_hours = leave.get('hours', 0)
            l_inits = leave.get('initials', '').strip().upper()
 
            if not l_type:
                continue
            # Bereavement can have no hours — skip only if no type and no relationship
            berv_rel = leave.get('bervRel', '').strip()
            if l_type != 'Bereavement' and not l_hours:
                continue
            if l_type == 'Bereavement' and not l_hours and not berv_rel:
                continue
 
            if week_start:
                from datetime import timedelta
                l_date = week_start + timedelta(days=leave_day_offsets[i])
                ws.cell(row=leave_row, column=1).value = l_date
                ws.cell(row=leave_row, column=1).number_format = 'M/D/YYYY'
 
            col_map = {'PTO': 2, 'Comp': 3, 'Sick': 4}
            if l_type in col_map:
                ws.cell(row=leave_row, column=col_map[l_type]).value = float(l_hours) if l_hours else None
            elif l_type == 'Bereavement':
                # Bereavement: write relationship text to column E, hours optional
                berv_rel = leave.get('bervRel', '').strip()
                if berv_rel:
                    ws.cell(row=leave_row, column=5).value = berv_rel
                if l_hours:
                    ws.cell(row=leave_row, column=4).value = float(l_hours)
 
            if l_inits:
                ws.cell(row=leave_row, column=11).value = l_inits
 
            leave_row += 1
 
        # ── Save to bytes and return ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
 
        sheet_num = str(ws['J4'].value or 'Timesheet').replace(' ', '_')
        filename  = f'{sheet_num}_filled.xlsx'
 
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
 
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500
 
if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
 
